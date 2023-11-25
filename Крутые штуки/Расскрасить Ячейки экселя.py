import pandas as pd

file_name = 'G:\Python\PycharmProjects\magazin\output1.xlsx'

import pandas as pd
import openpyxl

# загружаем данные из файла Excel
df = pd.read_excel(file_name)

# выбираем столбцы 'имя', 'дата', 'адрес' и проверяем, есть ли в них пустые ячейки
mask = df[['имя', 'дата', 'адрес']].isnull()

# заменяем пустые ячейки на значение "требует проверки"
df.loc[mask.any(axis=1), ['имя', 'дата', 'адрес']] = 'требует проверки'

# сохраняем измененный датафрейм в Excel-файл
df.to_excel('имя_файла.xlsx', index=False)

# загружаем файл Excel с помощью openpyxl
workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.active

# получаем размеры таблицы
max_row = worksheet.max_row
max_column = worksheet.max_column

# проходимся по каждой ячейке и покрашиваем ячейки, содержащие "требует проверки", в оранжевый цвет
for row in range(1, max_row+1):
    for column in range(1, max_column+1):
        cell = worksheet.cell(row=row, column=column)
        if cell.value == 'требует проверки':
            cell.fill = openpyxl.styles.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

# сохраняем изменения в файле Excel
workbook.save(file_name)